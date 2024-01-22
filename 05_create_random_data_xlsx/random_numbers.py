import os
import random
from openpyxl import Workbook

# Function to generate a random price between 1 and 300 as a float


def generate_random_price():
    return round(random.uniform(1, 300), 2)


# List of example company names
company_names = ["ABC Inc.", "XYZ Corporation", "Tech Solutions Ltd",
                 "Global Innovations", "Alpha Enterprises", "Beta Industries"]

# Generate 45 rows with data
data = []
for i in range(1, 46):
    row_number = i
    inventory = random.randint(10, 100)
    price = generate_random_price()
    company_name = random.choice(company_names)
    data.append((row_number, inventory, price, company_name))

# Create a new Excel workbook
workbook = Workbook()

# Select the active sheet
sheet = workbook.active

# Write the headers
headers = ["Row Number", "Inventory", "Price", "Company Name"]
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

# Write the data to columns A, B, C, and D
for row_num, row_data in enumerate(data, start=2):
    for col_num, value in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to a file with the specified filename
filename = "current_dir, random_data_with_companies_and_inventory.xlsx"
workbook.save(filename)

# Print the absolute path of the saved file
absolute_path = os.path.abspath(filename)
print(f"File saved at: {absolute_path}")
